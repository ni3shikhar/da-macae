"""
Security Assessment Excel Report Generator
==========================================
Generates comprehensive Excel reports for MCSB v2 security assessments.

Report Structure:
  - Executive Summary sheet
  - Domain Summary sheet  
  - Per-domain detail sheets (12 sheets for NS, IM, DP, PA, AM, LT, IR, PV, ES, BR, DS, GS)
"""

from __future__ import annotations

import asyncio
import io
import json
import os
from datetime import datetime
from typing import Any

# openpyxl for Excel generation
from openpyxl import Workbook
from openpyxl.styles import (
    Font, 
    Fill, 
    PatternFill, 
    Border, 
    Side, 
    Alignment,
    NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference


# ══════════════════════════════════════════════════════════════════════
# Constants
# ══════════════════════════════════════════════════════════════════════

DOMAIN_INFO = {
    "NS": {"name": "Network Security", "order": 1},
    "IM": {"name": "Identity Management", "order": 2},
    "DP": {"name": "Data Protection", "order": 3},
    "PA": {"name": "Privileged Access", "order": 4},
    "AM": {"name": "Asset Management", "order": 5},
    "LT": {"name": "Logging and Threat Detection", "order": 6},
    "IR": {"name": "Incident Response", "order": 7},
    "PV": {"name": "Posture and Vulnerability Management", "order": 8},
    "ES": {"name": "Endpoint Security", "order": 9},
    "BR": {"name": "Backup and Recovery", "order": 10},
    "DS": {"name": "DevOps Security", "order": 11},
    "GS": {"name": "Governance and Strategy", "order": 12},
}

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MANUAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LOW_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
INFO_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _get_status_fill(status: str) -> PatternFill:
    """Get cell fill color based on status."""
    status_upper = status.upper()
    if status_upper == "PASS":
        return PASS_FILL
    elif status_upper == "FAIL":
        return FAIL_FILL
    elif status_upper == "MANUAL_REVIEW":
        return MANUAL_FILL
    return PatternFill()


def _get_severity_fill(severity: str) -> PatternFill:
    """Get cell fill color based on severity."""
    severity_lower = severity.lower()
    if severity_lower == "critical":
        return CRITICAL_FILL
    elif severity_lower == "high":
        return HIGH_FILL
    elif severity_lower == "medium":
        return MEDIUM_FILL
    elif severity_lower == "low":
        return LOW_FILL
    elif severity_lower == "info":
        return INFO_FILL
    return PatternFill()


def _auto_width(ws, min_width: int = 10, max_width: int = 50):
    """Auto-adjust column widths based on content."""
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column].width = adjusted_width


def generate_security_report(
    findings: list[dict],
    subscription_id: str = "",
    subscription_name: str = "",
) -> bytes:
    """
    Generate an Excel report from security assessment findings.
    
    Args:
        findings: List of finding dictionaries from domain assessments.
        subscription_id: Azure subscription ID.
        subscription_name: Azure subscription display name.
    
    Returns:
        Excel file as bytes.
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # ── Organize findings by domain ─────────────────────────────────
    findings_by_domain: dict[str, list[dict]] = {d: [] for d in DOMAIN_INFO.keys()}
    
    for finding in findings:
        ctrl_id = finding.get("control_id", "")
        domain = ctrl_id.split("-")[0] if "-" in ctrl_id else ""
        if domain in findings_by_domain:
            findings_by_domain[domain].append(finding)
    
    # ── Calculate statistics ────────────────────────────────────────
    total_findings = len(findings)
    pass_count = sum(1 for f in findings if f.get("status", "").upper() == "PASS")
    fail_count = sum(1 for f in findings if f.get("status", "").upper() == "FAIL")
    manual_count = sum(1 for f in findings if f.get("status", "").upper() == "MANUAL_REVIEW")
    
    critical_count = sum(1 for f in findings if f.get("severity", "").lower() == "critical" and f.get("status", "").upper() == "FAIL")
    high_count = sum(1 for f in findings if f.get("severity", "").lower() == "high" and f.get("status", "").upper() == "FAIL")
    
    compliance_pct = (pass_count / total_findings * 100) if total_findings > 0 else 0
    
    # ── Executive Summary Sheet ─────────────────────────────────────
    ws_exec = wb.create_sheet("Executive Summary")
    
    # Title
    ws_exec.merge_cells("A1:F1")
    ws_exec["A1"] = "MCSB v2 Security Assessment Report"
    ws_exec["A1"].font = Font(bold=True, size=18, color="1F4E79")
    ws_exec["A1"].alignment = Alignment(horizontal="center")
    
    # Metadata
    ws_exec["A3"] = "Assessment Date:"
    ws_exec["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws_exec["A4"] = "Subscription ID:"
    ws_exec["B4"] = subscription_id or "N/A"
    ws_exec["A5"] = "Subscription Name:"
    ws_exec["B5"] = subscription_name or "N/A"
    ws_exec["A6"] = "Benchmark:"
    ws_exec["B6"] = "Microsoft Cloud Security Benchmark v2"
    
    for row in range(3, 7):
        ws_exec[f"A{row}"].font = Font(bold=True)
    
    # Overall Score
    ws_exec["A8"] = "Overall Compliance Score"
    ws_exec["A8"].font = Font(bold=True, size=14)
    ws_exec["B8"] = f"{compliance_pct:.1f}%"
    ws_exec["B8"].font = Font(bold=True, size=24, color="1F4E79")
    
    # Summary Statistics
    ws_exec["A10"] = "Summary Statistics"
    ws_exec["A10"].font = Font(bold=True, size=14)
    
    stats_headers = ["Metric", "Count", "Percentage"]
    for col, header in enumerate(stats_headers, 1):
        cell = ws_exec.cell(row=11, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    stats_data = [
        ("Total Findings", total_findings, "100%"),
        ("Passed", pass_count, f"{pass_count/total_findings*100:.1f}%" if total_findings else "0%"),
        ("Failed", fail_count, f"{fail_count/total_findings*100:.1f}%" if total_findings else "0%"),
        ("Manual Review", manual_count, f"{manual_count/total_findings*100:.1f}%" if total_findings else "0%"),
    ]
    
    for row_idx, (metric, count, pct) in enumerate(stats_data, 12):
        ws_exec.cell(row=row_idx, column=1, value=metric).border = THIN_BORDER
        ws_exec.cell(row=row_idx, column=2, value=count).border = THIN_BORDER
        ws_exec.cell(row=row_idx, column=3, value=pct).border = THIN_BORDER
        
        if metric == "Passed":
            ws_exec.cell(row=row_idx, column=2).fill = PASS_FILL
        elif metric == "Failed":
            ws_exec.cell(row=row_idx, column=2).fill = FAIL_FILL
        elif metric == "Manual Review":
            ws_exec.cell(row=row_idx, column=2).fill = MANUAL_FILL
    
    # Critical/High Findings
    ws_exec["A17"] = "Urgent Findings"
    ws_exec["A17"].font = Font(bold=True, size=14)
    
    ws_exec["A18"] = "Critical Severity Failures:"
    ws_exec["B18"] = critical_count
    ws_exec["B18"].fill = CRITICAL_FILL if critical_count > 0 else PatternFill()
    ws_exec["B18"].font = Font(bold=True, color="FFFFFF" if critical_count > 0 else "000000")
    
    ws_exec["A19"] = "High Severity Failures:"
    ws_exec["B19"] = high_count
    ws_exec["B19"].fill = HIGH_FILL if high_count > 0 else PatternFill()
    ws_exec["B19"].font = Font(bold=True)
    
    # Top 5 Critical Findings
    ws_exec["A21"] = "Top Critical Findings"
    ws_exec["A21"].font = Font(bold=True, size=14)
    
    critical_findings = [
        f for f in findings 
        if f.get("status", "").upper() == "FAIL" 
        and f.get("severity", "").lower() in ("critical", "high")
    ]
    critical_findings.sort(key=lambda x: (0 if x.get("severity", "").lower() == "critical" else 1, x.get("control_id", "")))
    
    crit_headers = ["Control ID", "Title", "Severity", "Resource", "Finding"]
    for col, header in enumerate(crit_headers, 1):
        cell = ws_exec.cell(row=22, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    for row_idx, finding in enumerate(critical_findings[:10], 23):
        ws_exec.cell(row=row_idx, column=1, value=finding.get("control_id", "")).border = THIN_BORDER
        ws_exec.cell(row=row_idx, column=2, value=finding.get("title", "")).border = THIN_BORDER
        sev_cell = ws_exec.cell(row=row_idx, column=3, value=finding.get("severity", ""))
        sev_cell.border = THIN_BORDER
        sev_cell.fill = _get_severity_fill(finding.get("severity", ""))
        ws_exec.cell(row=row_idx, column=4, value=finding.get("resource_name", "")).border = THIN_BORDER
        ws_exec.cell(row=row_idx, column=5, value=finding.get("finding", "")[:100]).border = THIN_BORDER
    
    _auto_width(ws_exec)
    
    # ── Domain Summary Sheet ────────────────────────────────────────
    ws_domain = wb.create_sheet("Domain Summary")
    
    ws_domain["A1"] = "Domain Summary"
    ws_domain["A1"].font = Font(bold=True, size=16, color="1F4E79")
    
    domain_headers = ["Domain ID", "Domain Name", "Total", "Pass", "Fail", "Manual", "Compliance %"]
    for col, header in enumerate(domain_headers, 1):
        cell = ws_domain.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    row_idx = 4
    for domain_id in sorted(DOMAIN_INFO.keys(), key=lambda d: DOMAIN_INFO[d]["order"]):
        domain_findings = findings_by_domain[domain_id]
        domain_total = len(domain_findings)
        domain_pass = sum(1 for f in domain_findings if f.get("status", "").upper() == "PASS")
        domain_fail = sum(1 for f in domain_findings if f.get("status", "").upper() == "FAIL")
        domain_manual = sum(1 for f in domain_findings if f.get("status", "").upper() == "MANUAL_REVIEW")
        domain_compliance = (domain_pass / domain_total * 100) if domain_total > 0 else 0
        
        ws_domain.cell(row=row_idx, column=1, value=domain_id).border = THIN_BORDER
        ws_domain.cell(row=row_idx, column=2, value=DOMAIN_INFO[domain_id]["name"]).border = THIN_BORDER
        ws_domain.cell(row=row_idx, column=3, value=domain_total).border = THIN_BORDER
        
        pass_cell = ws_domain.cell(row=row_idx, column=4, value=domain_pass)
        pass_cell.border = THIN_BORDER
        if domain_pass > 0:
            pass_cell.fill = PASS_FILL
        
        fail_cell = ws_domain.cell(row=row_idx, column=5, value=domain_fail)
        fail_cell.border = THIN_BORDER
        if domain_fail > 0:
            fail_cell.fill = FAIL_FILL
        
        manual_cell = ws_domain.cell(row=row_idx, column=6, value=domain_manual)
        manual_cell.border = THIN_BORDER
        if domain_manual > 0:
            manual_cell.fill = MANUAL_FILL
        
        compliance_cell = ws_domain.cell(row=row_idx, column=7, value=f"{domain_compliance:.1f}%")
        compliance_cell.border = THIN_BORDER
        
        row_idx += 1
    
    _auto_width(ws_domain)
    
    # ── Per-Domain Detail Sheets ────────────────────────────────────
    detail_headers = [
        "Control ID", "Title", "Status", "Severity", "Resource Name", 
        "Resource Type", "Resource Group", "Current Value", "Expected Value",
        "Finding", "Recommendation", "Rationale"
    ]
    
    for domain_id in sorted(DOMAIN_INFO.keys(), key=lambda d: DOMAIN_INFO[d]["order"]):
        domain_name = DOMAIN_INFO[domain_id]["name"]
        sheet_name = f"{domain_id} - {domain_name}"[:31]  # Excel sheet name limit
        
        ws = wb.create_sheet(sheet_name)
        
        # Title
        ws.merge_cells("A1:L1")
        ws["A1"] = f"{domain_id}: {domain_name}"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        
        # Headers
        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        
        # Data rows
        domain_findings = findings_by_domain[domain_id]
        domain_findings.sort(key=lambda f: f.get("control_id", ""))
        
        for row_idx, finding in enumerate(domain_findings, 4):
            ws.cell(row=row_idx, column=1, value=finding.get("control_id", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=finding.get("title", "")).border = THIN_BORDER
            
            status_cell = ws.cell(row=row_idx, column=3, value=finding.get("status", ""))
            status_cell.border = THIN_BORDER
            status_cell.fill = _get_status_fill(finding.get("status", ""))
            
            sev_cell = ws.cell(row=row_idx, column=4, value=finding.get("severity", ""))
            sev_cell.border = THIN_BORDER
            sev_cell.fill = _get_severity_fill(finding.get("severity", ""))
            
            ws.cell(row=row_idx, column=5, value=finding.get("resource_name", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=6, value=finding.get("resource_type", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=7, value=finding.get("resource_group", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=8, value=finding.get("current_value", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=9, value=finding.get("expected_value", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=10, value=finding.get("finding", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=11, value=finding.get("recommendation", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=12, value=finding.get("rationale", "")).border = THIN_BORDER
        
        # Auto-filter
        if len(domain_findings) > 0:
            ws.auto_filter.ref = f"A3:L{3 + len(domain_findings)}"
        
        # Freeze panes
        ws.freeze_panes = "A4"
        
        # Set row height for readability
        for row in range(4, 4 + len(domain_findings)):
            ws.row_dimensions[row].height = 30
        
        _auto_width(ws)
    
    # ── Save to bytes ───────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


async def sec_generate_excel_report(
    findings_json: str,
    subscription_id: str = "",
    subscription_name: str = "",
) -> str:
    """
    Generate an Excel security assessment report from findings JSON.
    
    Args:
        findings_json: JSON string containing array of finding objects.
        subscription_id: Optional Azure subscription ID for the report.
        subscription_name: Optional Azure subscription name for the report.
    
    Returns:
        JSON with status and base64-encoded Excel file.
    """
    import base64
    
    loop = asyncio.get_running_loop()
    
    def _generate():
        try:
            # Parse findings
            if isinstance(findings_json, str):
                data = json.loads(findings_json)
            else:
                data = findings_json
            
            # Extract findings array
            if isinstance(data, list):
                findings = data
            elif isinstance(data, dict) and "findings" in data:
                findings = data["findings"]
            else:
                findings = []
            
            # Generate Excel
            excel_bytes = generate_security_report(
                findings=findings,
                subscription_id=subscription_id,
                subscription_name=subscription_name,
            )
            
            # Encode as base64
            excel_b64 = base64.b64encode(excel_bytes).decode("utf-8")
            
            # Calculate summary stats
            total = len(findings)
            passed = sum(1 for f in findings if f.get("status", "").upper() == "PASS")
            failed = sum(1 for f in findings if f.get("status", "").upper() == "FAIL")
            manual = sum(1 for f in findings if f.get("status", "").upper() == "MANUAL_REVIEW")
            
            return {
                "status": "success",
                "filename": f"mcsb-v2-assessment-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.xlsx",
                "excel_base64": excel_b64,
                "size_bytes": len(excel_bytes),
                "summary": {
                    "total_findings": total,
                    "passed": passed,
                    "failed": failed,
                    "manual_review": manual,
                    "compliance_percentage": (passed / total * 100) if total > 0 else 0,
                },
            }
        except Exception as e:
            return {"status": "error", "error": str(e)}
    
    result = await loop.run_in_executor(None, _generate)
    return json.dumps(result, default=str)


def register_excel_tools(mcp_server):
    """
    Register Excel report generation tools with the MCP server.
    
    Args:
        mcp_server: The FastMCP server instance to register tools on.
    """
    mcp_server.tool()(sec_generate_excel_report)
    print("[INFO] Security Excel report tools registered successfully")
